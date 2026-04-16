"""
Phase 1 — Stock Portfolio Performance Tracker
Pull historical price data via yfinance and export to Excel + CSV.

Usage:
    pip install yfinance openpyxl pandas
    python phase1_fetch_stock_data.py
"""

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

# ─────────────────────────────────────────────
# CONFIG — edit these to match your portfolio
# ─────────────────────────────────────────────
TICKERS = ["AAPL", "MSFT", "GOOGL", "AMZN", "NVDA"]
START_DATE = "2022-01-01"
END_DATE   = datetime.today().strftime("%Y-%m-%d")
OUTPUT_DIR = "output"
# ─────────────────────────────────────────────


def fetch_price_data(tickers: list[str], start: str, end: str) -> pd.DataFrame:
    """Download adjusted close prices for all tickers."""
    print(f"Fetching data for: {', '.join(tickers)}")
    raw = yf.download(tickers, start=start, end=end, auto_adjust=True, progress=False)

    # yfinance returns MultiIndex columns when multiple tickers
    if isinstance(raw.columns, pd.MultiIndex):
        prices = raw["Close"].copy()
    else:
        prices = raw[["Close"]].copy()
        prices.columns = tickers

    prices.index.name = "Date"
    prices = prices.round(4)
    print(f"  Retrieved {len(prices)} trading days ({start} → {end})")
    return prices


def compute_daily_returns(prices: pd.DataFrame) -> pd.DataFrame:
    """Simple daily % return for each ticker."""
    returns = prices.pct_change().round(6)
    returns.index.name = "Date"
    return returns


def compute_summary_stats(prices: pd.DataFrame) -> pd.DataFrame:
    """Per-ticker summary: start price, end price, total return, vol, max drawdown."""
    records = []
    for ticker in prices.columns:
        s = prices[ticker].dropna()
        if s.empty:
            continue
        start_px = s.iloc[0]
        end_px   = s.iloc[-1]
        total_ret = (end_px - start_px) / start_px

        daily_ret = s.pct_change().dropna()
        annual_vol = daily_ret.std() * (252 ** 0.5)

        # Max drawdown
        rolling_max = s.cummax()
        drawdown    = (s - rolling_max) / rolling_max
        max_dd      = drawdown.min()

        records.append({
            "Ticker":          ticker,
            "Start Date":      s.index[0].strftime("%Y-%m-%d"),
            "End Date":        s.index[-1].strftime("%Y-%m-%d"),
            "Start Price ($)": round(start_px, 2),
            "End Price ($)":   round(end_px, 2),
            "Total Return":    round(total_ret, 4),
            "Annual Volatility": round(annual_vol, 4),
            "Max Drawdown":    round(max_dd, 4),
            "Trading Days":    len(s),
        })

    return pd.DataFrame(records)


def style_header_row(ws, row: int, col_start: int, col_end: int, fill_hex="1F4E79"):
    """Apply dark-blue header styling to a row range."""
    fill   = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        bottom=Side(style="medium", color="FFFFFF"),
    )
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = font
        cell.fill   = fill
        cell.alignment = align
        cell.border = border


def set_col_widths(ws, headers, min_width=10, max_width=30):
    """Set column widths based on header lengths only — avoids iterating all cells."""
    for i, h in enumerate(headers, start=1):
        width = min(max(len(str(h)) + 4, min_width), max_width)
        ws.column_dimensions[get_column_letter(i)].width = width


def write_excel(prices: pd.DataFrame, returns: pd.DataFrame, summary: pd.DataFrame, path: str):
    print("  Building Excel workbook...")
    wb = Workbook()

    # ── Sheet 1: Summary Stats ───────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.row_dimensions[1].height = 20

    title_cell = ws_sum["A1"]
    title_cell.value = "Stock Portfolio — Phase 1 Summary Statistics"
    title_cell.font  = Font(bold=True, size=13, name="Arial", color="1F4E79")
    ws_sum.merge_cells(f"A1:{get_column_letter(len(summary.columns))}1")
    ws_sum["A2"].value = f"Generated: {datetime.today().strftime('%B %d, %Y')}"
    ws_sum["A2"].font  = Font(italic=True, size=9, name="Arial", color="595959")

    headers = list(summary.columns)
    for col_idx, h in enumerate(headers, start=1):
        ws_sum.cell(row=3, column=col_idx, value=h)
    style_header_row(ws_sum, 3, 1, len(headers))

    pct_cols   = {"Total Return", "Annual Volatility", "Max Drawdown"}
    price_cols = {"Start Price ($)", "End Price ($)"}
    zebra_fill = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")

    for r_idx, row_data in enumerate(summary.itertuples(index=False), start=4):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws_sum.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")
            col_name = headers[c_idx - 1]
            if col_name in pct_cols:
                cell.number_format = "0.00%"
            elif col_name in price_cols:
                cell.number_format = "$#,##0.00"
            if r_idx % 2 == 0:
                cell.fill = zebra_fill

    set_col_widths(ws_sum, headers)

    # ── Sheet 2: Daily Prices ────────────────────────────────────────────────
    print("  Writing Daily Prices sheet...")
    ws_px = wb.create_sheet("Daily Prices")
    px_headers = ["Date"] + list(prices.columns)
    for c, h in enumerate(px_headers, 1):
        ws_px.cell(row=1, column=c, value=h)
    style_header_row(ws_px, 1, 1, len(px_headers))

    grey_fill = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
    arial10   = Font(name="Arial", size=10)

    for r, (date, row_data) in enumerate(prices.iterrows(), start=2):
        ws_px.cell(row=r, column=1, value=date.strftime("%Y-%m-%d"))
        for c, val in enumerate(row_data, start=2):
            cell = ws_px.cell(row=r, column=c, value=round(float(val), 4) if pd.notna(val) else None)
            cell.number_format = "$#,##0.0000"
            cell.font = arial10
            if r % 2 == 0:
                cell.fill = grey_fill

    set_col_widths(ws_px, px_headers)

    # ── Sheet 3: Daily Returns ───────────────────────────────────────────────
    print("  Writing Daily Returns sheet...")
    ws_ret = wb.create_sheet("Daily Returns")
    ret_headers = ["Date"] + list(returns.columns)
    for c, h in enumerate(ret_headers, 1):
        ws_ret.cell(row=1, column=c, value=h)
    style_header_row(ws_ret, 1, 1, len(ret_headers), fill_hex="1A3C5E")

    pos_fill = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")
    neg_fill = PatternFill("solid", start_color="FFEBEE", end_color="FFEBEE")

    for r, (date, row_data) in enumerate(returns.dropna().iterrows(), start=2):
        ws_ret.cell(row=r, column=1, value=date.strftime("%Y-%m-%d"))
        for c, val in enumerate(row_data, start=2):
            cell = ws_ret.cell(row=r, column=c, value=round(float(val), 6) if pd.notna(val) else None)
            cell.number_format = "0.0000%"
            cell.font = arial10
            if pd.notna(val):
                cell.fill = pos_fill if val >= 0 else neg_fill

    set_col_widths(ws_ret, ret_headers)

    print("  Saving file (this may take ~10–20s for 1000+ rows)...")
    wb.save(path)
    print(f"  Excel saved → {path}")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    prices  = fetch_price_data(TICKERS, START_DATE, END_DATE)
    returns = compute_daily_returns(prices)
    summary = compute_summary_stats(prices)

    # ── CSV exports (for SQL ingestion in Phase 2) ───────────────────────────
    prices_long = prices.reset_index().melt(id_vars="Date", var_name="Ticker", value_name="Close_Price")
    prices_long["Date"] = prices_long["Date"].dt.strftime("%Y-%m-%d")
    prices_long = prices_long.sort_values(["Ticker", "Date"]).reset_index(drop=True)

    returns_long = returns.reset_index().melt(id_vars="Date", var_name="Ticker", value_name="Daily_Return")
    returns_long["Date"] = returns_long["Date"].dt.strftime("%Y-%m-%d")
    returns_long = returns_long.sort_values(["Ticker", "Date"]).reset_index(drop=True)

    csv_prices  = os.path.join(OUTPUT_DIR, "stock_prices.csv")
    csv_returns = os.path.join(OUTPUT_DIR, "stock_returns.csv")
    csv_summary = os.path.join(OUTPUT_DIR, "stock_summary.csv")

    prices_long.to_csv(csv_prices,  index=False)
    returns_long.to_csv(csv_returns, index=False)
    summary.to_csv(csv_summary, index=False)
    print(f"  CSVs saved → {OUTPUT_DIR}/")

    # ── Excel workbook ───────────────────────────────────────────────────────
    excel_path = os.path.join(OUTPUT_DIR, "portfolio_phase1.xlsx")
    write_excel(prices, returns, summary, excel_path)

    print("\n✅ Phase 1 complete. Output files:")
    for f in [csv_prices, csv_returns, csv_summary, excel_path]:
        print(f"   {f}")
    print("\nNext → Phase 2: Load CSVs into SQL (CREATE TABLE + INSERT + analytics queries)")


if __name__ == "__main__":
    main()